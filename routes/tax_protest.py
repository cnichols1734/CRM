"""
Tax Protest blueprint.
Lets agents search their CRM contacts, locate properties in county tax data,
extract subdivisions via LLM, find lower-value comparables, and export results.
"""

import csv
import logging
import os
import re
import time
from functools import lru_cache
from io import BytesIO, StringIO

from flask import (
    Blueprint,
    render_template,
    request,
    jsonify,
    Response,
    abort,
    send_file,
    session,
)
from flask_login import login_required, current_user
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill
from PIL import Image as PILImage, ImageDraw, ImageFont
from werkzeug.exceptions import HTTPException

try:
    import psutil
except ImportError:  # pragma: no cover - psutil is available in production
    psutil = None

from forms import ContactForm
from models import db, Contact, ContactGroup
from feature_flags import feature_required
from services.tenant_service import org_query, can_view_all_org_data
from services.tax_protest_service import (
    find_property_in_tax_data,
    extract_chambers_subdivision,
    extract_subdivision_llm,
    find_comparables,
    get_neighborhood_name,
    get_subdivision_stats,
    cache_search_result,
    get_cached_search_result,
    get_main_property_by_id,
    _is_valid_subdivision,
    build_chambers_subdivision_match_terms,
)

tax_protest_bp = Blueprint("tax_protest", __name__, url_prefix="/tax-protest")
logger = logging.getLogger(__name__)

COUNTY_LABELS = {
    "chambers": "Chambers",
    "hcad": "Harris",
    "liberty": "Liberty",
    "fort_bend": "Fort Bend",
}
EXPORT_XLSX_MIMETYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)
EXPORT_FONT_PATH = "/System/Library/Fonts/Helvetica.ttc"
SEARCH_COMPARABLE_LIMIT = 250


def _elapsed_ms(started_at):
    return round((time.perf_counter() - started_at) * 1000, 1)


def _current_rss_mb():
    if psutil is None:
        return None
    try:
        process = psutil.Process(os.getpid())
        return round(process.memory_info().rss / 1024 / 1024, 1)
    except Exception:
        return None


def _log_tax_event(event, **fields):
    payload = " ".join(
        f"{key}={value}"
        for key, value in fields.items()
        if value is not None and value != ""
    )
    logger.info("tax_protest_%s %s", event, payload)


def _coerce_positive_number(value):
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    return number if number > 0 else None


def _authorized_contact(contact_id):
    """Load a contact with org + ownership checks. Returns contact or aborts 403."""
    contact = org_query(Contact).filter_by(id=contact_id).first()
    if not contact:
        abort(404, description="Contact not found")
    if not can_view_all_org_data() and contact.user_id != current_user.id:
        abort(403, description="You can only search your own contacts")
    return contact


def _safe_filename(address, suffix):
    addr = address or "unknown"
    base = re.sub(r"[^a-zA-Z0-9]+", "_", addr).strip("_") or "unknown"
    return f"{base}{suffix}"


def _contact_search_payload(contact):
    addr_parts = [
        part
        for part in [contact.street_address, contact.city, contact.state, contact.zip_code]
        if part
    ]
    return {
        "id": contact.id,
        "name": f"{contact.first_name} {contact.last_name}",
        "address": ", ".join(addr_parts),
        "street_address": contact.street_address,
        "city": contact.city,
        "state": contact.state,
        "zip_code": contact.zip_code,
        "has_address": bool(contact.street_address and contact.street_address.strip()),
    }


def _format_axis_value(value):
    if value is None:
        return ""
    if abs(value) >= 1000000:
        return f"${value / 1000000:.1f}M"
    return f"${round(value / 1000):.0f}k"


def _format_currency(value):
    if value in (None, ""):
        return "—"
    return f"${value:,.0f}"


def _ordinal_suffix(value):
    value = abs(int(value))
    if 10 <= value % 100 <= 20:
        return "th"
    return {1: "st", 2: "nd", 3: "rd"}.get(value % 10, "th")


@lru_cache(maxsize=None)
def _load_export_font(size):
    try:
        return ImageFont.truetype(EXPORT_FONT_PATH, size)
    except OSError:
        return ImageFont.load_default()


def _text_size(draw, text, font):
    left, top, right, bottom = draw.textbbox((0, 0), text, font=font)
    return right - left, bottom - top


def _wrap_text(draw, text, font, max_width):
    words = text.split()
    if not words:
        return [""]

    lines = []
    current = words[0]
    for word in words[1:]:
        candidate = f"{current} {word}"
        if _text_size(draw, candidate, font)[0] <= max_width:
            current = candidate
        else:
            lines.append(current)
            current = word
    lines.append(current)
    return lines


def _draw_dashed_line(draw, x, y1, y2, dash_length=10, gap=7, fill="#f59e0b", width=3):
    y = y1
    while y < y2:
        dash_end = min(y + dash_length, y2)
        draw.line((x, y, x, dash_end), fill=fill, width=width)
        y = dash_end + gap


def _has_positive_market_value(prop):
    value = prop.get("market_value")
    try:
        return float(value) > 0
    except (TypeError, ValueError):
        return False


def _build_export_rows(main_property, comparables, subdivision, county):
    rows = []

    def build_row(prop, row_type):
        if not _has_positive_market_value(prop):
            return
        rows.append(
            [
                row_type,
                prop.get("full_address", ""),
                prop.get("city", ""),
                prop.get("zip", ""),
                prop.get("market_value", ""),
                prop.get("sq_ft", ""),
                prop.get("acreage", ""),
                subdivision,
                prop.get("legal1", ""),
                prop.get("account", ""),
                county,
            ]
        )

    build_row(main_property, "Subject Property")
    for comp in comparables:
        build_row(comp, "Comparable")
    return rows


def _load_cached_export_data():
    cached = get_cached_search_result()
    if not cached:
        abort(400, description="No search results to download. Run a search first.")

    contact = _authorized_contact(cached["contact_id"])
    main_property = get_main_property_by_id(
        cached["main_property_id"], cached["source"]
    )
    if not main_property:
        abort(404, description="Main property no longer found in tax data")

    comparables = find_comparables(
        cached["subdivision"],
        cached["zip_code"],
        main_property["market_value"],
        cached["source"],
        main_sq_ft=cached.get("main_sq_ft"),
        fuzzy_subdivision=cached.get("fuzzy_subdivision", False),
        subdivision_code=cached.get("subdivision_code"),
        main_acreage=cached.get("main_acreage"),
        subdivision_match_terms=cached.get("subdivision_match_terms"),
    )
    subdivision_stats = get_subdivision_stats(
        cached["subdivision"],
        cached["zip_code"],
        main_property["market_value"],
        cached["source"],
        fuzzy_subdivision=cached.get("fuzzy_subdivision", False),
        subdivision_code=cached.get("subdivision_code"),
        subdivision_match_terms=cached.get("subdivision_match_terms"),
    )

    county = COUNTY_LABELS.get(cached["source"], cached["source"].title())
    subdivision = cached.get("subdivision", "")

    return {
        "cached": cached,
        "contact": contact,
        "main_property": main_property,
        "comparables": comparables,
        "subdivision_stats": subdivision_stats,
        "county": county,
        "subdivision": subdivision,
        "rows": _build_export_rows(main_property, comparables, subdivision, county),
    }


def _build_chart_image(stats, subdivision, subject_value):
    if (
        not stats
        or not stats.get("value_distribution")
        or stats.get("min_value") is None
        or stats.get("max_value") is None
    ):
        return None

    image_width = 1280
    image_height = 760
    image = PILImage.new("RGB", (image_width, image_height), "white")
    draw = ImageDraw.Draw(image)

    title_font = _load_export_font(34)
    subtitle_font = _load_export_font(20)
    label_font = _load_export_font(18)
    card_label_font = _load_export_font(15)
    value_font = _load_export_font(28)
    chart_title_font = _load_export_font(22)
    chart_label_font = _load_export_font(18)
    axis_font = _load_export_font(16)

    slate_950 = "#0f172a"
    slate_700 = "#334155"
    slate_500 = "#64748b"
    slate_300 = "#cbd5e1"
    slate_200 = "#e2e8f0"
    slate_50 = "#f8fafc"
    emerald = "#34d399"
    amber = "#fbbf24"
    amber_dark = "#b45309"
    orange = "#f97316"

    outer_box = (30, 30, image_width - 30, image_height - 30)
    draw.rounded_rectangle(outer_box, radius=28, fill="white", outline=slate_200, width=2)

    draw.text((72, 62), "Neighborhood Position", font=title_font, fill=slate_950)
    subtitle = f"{subdivision or 'Subdivision'} · {stats['total_homes']} homes"
    draw.text((72, 108), subtitle, font=subtitle_font, fill=slate_500)

    pct = round(stats.get("percentile") or 0)
    pill_text = f"{pct}{_ordinal_suffix(pct)} percentile"
    pill_fill = "#dcfce7" if pct >= 75 else "#fef3c7" if pct >= 50 else "#ffedd5" if pct >= 25 else "#e2e8f0"
    pill_text_fill = "#047857" if pct >= 75 else "#b45309" if pct >= 50 else orange if pct >= 25 else slate_700
    pill_width = _text_size(draw, pill_text, subtitle_font)[0] + 32
    pill_x = image_width - pill_width - 72
    pill_y = 68
    draw.rounded_rectangle(
        (pill_x, pill_y, pill_x + pill_width, pill_y + 40),
        radius=20,
        fill=pill_fill,
    )
    draw.text((pill_x + 16, pill_y + 8), pill_text, font=subtitle_font, fill=pill_text_fill)

    cards = [
        ("# of homes valued for less", str(stats.get("lower_values") or 0), "#ecfdf5", emerald),
        ("Subject Value", _format_currency(subject_value), "#fffbeb", amber),
        (
            "# of homes valued for more",
            str(stats.get("higher_values") or 0),
            slate_50,
            slate_300,
        ),
        (
            "Neighborhood Median",
            _format_currency(stats.get("median_value")),
            "#fff7ed",
            orange,
        ),
    ]
    card_y = 156
    card_w = 268
    card_h = 110
    card_gap = 20
    card_x = 72
    for label, value, fill, accent in cards:
        draw.rounded_rectangle(
            (card_x, card_y, card_x + card_w, card_y + card_h),
            radius=22,
            fill=fill,
            outline=slate_200,
            width=2,
        )
        draw.rounded_rectangle(
            (card_x + 18, card_y + 18, card_x + 34, card_y + 34),
            radius=5,
            fill=accent,
        )
        label_lines = _wrap_text(draw, label, card_label_font, card_w - 72)
        line_height = _text_size(draw, "Ag", card_label_font)[1]
        label_y = card_y + 12
        for line in label_lines:
            draw.text((card_x + 46, label_y), line, font=card_label_font, fill=slate_500)
            label_y += line_height + 2
        value_y = max(card_y + 56, label_y + 8)
        draw.text((card_x + 18, value_y), value, font=value_font, fill=slate_950)
        card_x += card_w + card_gap

    chart_panel = (72, 304, image_width - 72, image_height - 90)
    draw.rounded_rectangle(chart_panel, radius=24, fill="white", outline=slate_200, width=2)
    draw.text(
        (chart_panel[0] + 24, chart_panel[1] + 20),
        "Market value distribution",
        font=chart_title_font,
        fill=slate_950,
    )
    draw.text(
        (chart_panel[0] + 24, chart_panel[1] + 52),
        "Green buckets show homes valued for less, amber marks the subject bucket, slate shows homes valued for more.",
        font=label_font,
        fill=slate_500,
    )

    dist = stats["value_distribution"]
    min_value = stats["min_value"]
    max_value = stats["max_value"]
    chart_left = chart_panel[0] + 42
    chart_top = chart_panel[1] + 96
    chart_right = chart_panel[2] - 42
    chart_bottom = chart_panel[3] - 74
    chart_width = chart_right - chart_left
    chart_height = chart_bottom - chart_top

    draw.line((chart_left, chart_bottom, chart_right, chart_bottom), fill=slate_200, width=2)

    max_count = max(bucket["count"] for bucket in dist) or 1
    bucket_count = len(dist)
    value_range = max(max_value - min_value, 1)
    bucket_width = chart_width / max(bucket_count, 1)
    bucket_size = value_range / max(bucket_count, 1)
    subject_index = int((subject_value - min_value) / bucket_size) if bucket_size else 0
    subject_index = max(0, min(bucket_count - 1, subject_index))

    for index, bucket in enumerate(dist):
        bucket_x = chart_left + index * bucket_width
        gap = max(6, bucket_width * 0.11)
        bar_height = max(14, chart_height * bucket["count"] / max_count) if bucket["count"] else 0
        bar_top = chart_bottom - bar_height
        fill = emerald if index < subject_index else amber if index == subject_index else slate_300
        if bar_height:
            draw.rounded_rectangle(
                (
                    bucket_x + gap,
                    bar_top,
                    bucket_x + bucket_width - gap,
                    chart_bottom,
                ),
                radius=10,
                fill=fill,
            )
            count_text = str(bucket["count"])
            count_width, count_height = _text_size(draw, count_text, chart_label_font)
            draw.text(
                (
                    bucket_x + (bucket_width - count_width) / 2,
                    max(chart_top, bar_top - count_height - 10),
                ),
                count_text,
                font=chart_label_font,
                fill=slate_700,
            )

    marker_x = chart_left + chart_width * ((subject_value - min_value) / value_range)
    marker_x = max(chart_left + 6, min(chart_right - 6, marker_x))
    _draw_dashed_line(draw, marker_x, chart_top, chart_bottom + 6, fill=amber_dark, width=3)
    draw.ellipse(
        (marker_x - 7, chart_bottom - 7, marker_x + 7, chart_bottom + 7),
        fill=amber,
        outline=amber_dark,
        width=2,
    )
    subject_label = f"{_format_axis_value(subject_value)} subject"
    subject_label_width, _ = _text_size(draw, subject_label, chart_label_font)
    label_x = min(max(chart_left, marker_x - subject_label_width / 2), chart_right - subject_label_width)
    draw.text(
        (label_x, chart_bottom + 18),
        subject_label,
        font=chart_label_font,
        fill=amber_dark,
    )

    axis_labels = [
        (chart_left + bucket_width / 2, dist[0]["label"]),
        (chart_left + (bucket_count - 0.5) * bucket_width, _format_axis_value(max_value)),
    ]
    midpoint = bucket_count // 2
    if midpoint not in (0, bucket_count - 1):
        axis_labels.append((chart_left + (midpoint + 0.5) * bucket_width, dist[midpoint]["label"]))

    for center_x, label in axis_labels:
        label_width, _ = _text_size(draw, label, axis_font)
        draw.text(
            (center_x - label_width / 2, chart_bottom + 48),
            label,
            font=axis_font,
            fill=slate_500,
        )

    output = BytesIO()
    image.save(output, format="PNG")
    output.seek(0)
    return output


def _build_xlsx_report(export_data):
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Summary"
    summary.sheet_view.showGridLines = False

    comparables_sheet = workbook.create_sheet("Comparables")
    distribution_sheet = workbook.create_sheet("Distribution")

    title_font = Font(name="Aptos", size=18, bold=True, color="0F172A")
    label_font = Font(name="Aptos", size=11, bold=True, color="334155")
    value_font = Font(name="Aptos", size=11, color="0F172A")
    header_font = Font(name="Aptos", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="0F172A")
    section_fill = PatternFill("solid", fgColor="F8FAFC")
    label_alignment = Alignment(horizontal="left", vertical="center")
    value_alignment = Alignment(horizontal="right", vertical="center")

    for column, width in {
        "A": 28,
        "B": 34,
        "C": 28,
        "D": 18,
        "E": 20,
        "F": 18,
        "G": 18,
    }.items():
        summary.column_dimensions[column].width = width

    summary.merge_cells("A1:G1")
    summary["A1"] = "Tax Protest Neighborhood Report"
    summary["A1"].font = title_font
    summary["A1"].alignment = Alignment(horizontal="left", vertical="center")

    summary.merge_cells("A2:G2")
    summary["A2"] = export_data["main_property"].get("full_address", "")
    summary["A2"].font = Font(name="Aptos", size=12, color="64748B")
    summary["A2"].alignment = label_alignment

    summary["A4"] = "Subdivision"
    summary["B4"] = export_data["subdivision"]
    summary["C4"] = "County"
    summary["D4"] = export_data["county"]
    summary["E4"] = "Exported Rows"
    summary["F4"] = len(export_data["rows"])

    stats = export_data["subdivision_stats"] or {}
    pct = round(stats.get("percentile") or 0)
    summary["A5"] = "Subject Market Value"
    summary["B5"] = export_data["main_property"].get("market_value")
    summary["C5"] = "Neighborhood Median"
    summary["D5"] = stats.get("median_value")
    summary["E5"] = "Percentile"
    summary["F5"] = f"{pct}{_ordinal_suffix(pct)} percentile"

    summary["A6"] = "# of homes valued for less"
    summary["B6"] = stats.get("lower_values")
    summary["C6"] = "# of homes valued for more"
    summary["D6"] = stats.get("higher_values")
    summary["E6"] = "Total Homes"
    summary["F6"] = stats.get("total_homes")

    for cell in ("A4", "C4", "E4", "A5", "C5", "E5", "A6", "C6", "E6"):
        summary[cell].font = label_font
        summary[cell].alignment = label_alignment
    for cell in ("B4", "D4", "F4", "B5", "D5", "F5", "B6", "D6", "F6"):
        summary[cell].font = value_font
        summary[cell].fill = section_fill
        summary[cell].alignment = value_alignment

    for cell in ("B5", "D5"):
        summary[cell].number_format = "$#,##0"

    summary["A8"] = "Neighborhood Distribution"
    summary["A8"].font = label_font
    summary["A9"] = "This chart is embedded as an image so the exported workbook preserves the same visual story as the app."
    summary["A9"].font = Font(name="Aptos", size=10, color="64748B")

    chart_image = _build_chart_image(
        export_data["subdivision_stats"],
        export_data["subdivision"],
        export_data["main_property"].get("market_value"),
    )
    if chart_image:
        xl_image = XLImage(chart_image)
        xl_image.width = 1100
        xl_image.height = 654
        summary.add_image(xl_image, "A11")

    headers = [
        "Type",
        "Address",
        "City",
        "Zip",
        "Market Value",
        "Sq Ft",
        "Acreage",
        "Subdivision",
        "Legal Description",
        "Account",
        "County",
    ]
    comparables_sheet.append(headers)
    for row in export_data["rows"]:
        comparables_sheet.append(row)

    for cell in comparables_sheet[1]:
        cell.font = header_font
        cell.fill = header_fill

    comparables_sheet.freeze_panes = "A2"
    for column, width in enumerate((18, 34, 18, 12, 16, 12, 12, 22, 38, 16, 14), start=1):
        comparables_sheet.column_dimensions[chr(64 + column)].width = width

    for row in range(2, comparables_sheet.max_row + 1):
        comparables_sheet[f"E{row}"].number_format = "$#,##0"
        comparables_sheet[f"F{row}"].number_format = "#,##0"
        comparables_sheet[f"G{row}"].number_format = "0.00"

    distribution_sheet.append(["Bucket Label", "Home Count"])
    for cell in distribution_sheet[1]:
        cell.font = header_font
        cell.fill = header_fill
    for bucket in stats.get("value_distribution", []):
        distribution_sheet.append([bucket["label"], bucket["count"]])
    distribution_sheet["D1"] = "Subject Value"
    distribution_sheet["E1"] = export_data["main_property"].get("market_value")
    distribution_sheet["D2"] = "Minimum Value"
    distribution_sheet["E2"] = stats.get("min_value")
    distribution_sheet["D3"] = "Maximum Value"
    distribution_sheet["E3"] = stats.get("max_value")
    distribution_sheet["D4"] = "Percentile"
    distribution_sheet["E4"] = stats.get("percentile")
    distribution_sheet.column_dimensions["A"].width = 18
    distribution_sheet.column_dimensions["B"].width = 12
    distribution_sheet.column_dimensions["D"].width = 18
    distribution_sheet.column_dimensions["E"].width = 14
    for cell in ("E1", "E2", "E3"):
        distribution_sheet[cell].number_format = "$#,##0"

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output


@tax_protest_bp.route("/")
@login_required
@feature_required("TAX_PROTEST")
def index():
    """Main Tax Protest page."""
    contact_groups = (
        org_query(ContactGroup)
        .order_by(ContactGroup.sort_order.asc(), ContactGroup.name.asc())
        .all()
    )
    return render_template(
        "tax_protest/index.html",
        contact_groups=contact_groups,
        contact_form=ContactForm(),
    )


@tax_protest_bp.route("/search-contacts")
@login_required
@feature_required("TAX_PROTEST")
def search_contacts():
    """AJAX endpoint: search CRM contacts by name or address."""
    q = request.args.get("q", "").strip()
    if len(q) < 2:
        return jsonify([])

    query = org_query(Contact)

    if not can_view_all_org_data():
        query = query.filter(Contact.user_id == current_user.id)

    search_term = f"%{q}%"
    query = query.filter(
        db.or_(
            Contact.first_name.ilike(search_term),
            Contact.last_name.ilike(search_term),
            db.func.concat(Contact.first_name, " ", Contact.last_name).ilike(
                search_term
            ),
            Contact.street_address.ilike(search_term),
            Contact.city.ilike(search_term),
            Contact.zip_code.ilike(search_term),
        )
    ).order_by(Contact.first_name.asc(), Contact.last_name.asc()).limit(15)

    return jsonify([_contact_search_payload(contact) for contact in query.all()])


@tax_protest_bp.route("/search", methods=["POST"])
@login_required
@feature_required("TAX_PROTEST")
def search_property():
    """Search tax data for a contact's property and find comparables."""
    data = request.get_json()
    if not data or not data.get("contact_id"):
        return jsonify({"error": "contact_id required"}), 400

    contact = _authorized_contact(data["contact_id"])
    source = None
    route_started = time.perf_counter()
    log_context = {
        "contact_id": contact.id,
        "user_id": current_user.id,
        "org_id": current_user.organization_id,
    }
    _log_tax_event("search_started", **log_context, rss_mb=_current_rss_mb())

    if not contact.street_address:
        return jsonify({"error": "Contact has no street address on file"}), 400

    try:
        lookup_started = time.perf_counter()
        property_record, source = find_property_in_tax_data(
            contact.street_address, contact.city, contact.zip_code
        )
        lookup_ms = _elapsed_ms(lookup_started)
        _log_tax_event(
            "search_lookup_complete",
            **log_context,
            source=source,
            lookup_ms=lookup_ms,
            rss_mb=_current_rss_mb(),
        )

        if not property_record:
            return jsonify(
                {
                    "error": f'No property found matching "{contact.street_address}" in Chambers, Harris, Liberty, or Fort Bend County tax records'
                }
            ), 404

        market_value = _coerce_positive_number(property_record.get("market_value"))
        if market_value is None:
            _log_tax_event(
                "search_invalid_market_value",
                **log_context,
                source=source,
                property_id=property_record.get("id"),
                rss_mb=_current_rss_mb(),
            )
            return jsonify(
                {
                    "error": "Property has no market value available in tax data",
                    "main_property": property_record,
                    "source": source,
                }
            ), 422

        zip_code = property_record.get("zip")
        neighborhood_code = property_record.get("neighborhood_code")
        subdivision_code = property_record.get("subdivision_code")
        main_sq_ft = property_record.get("sq_ft")
        main_acreage = property_record.get("acreage")
        subdivision = None
        fuzzy = False
        subdivision_match_terms = None
        llm_ms = 0.0

        if source == "hcad":
            lgl_2 = property_record.get("legal2") or ""
            if _is_valid_subdivision(lgl_2):
                subdivision = lgl_2.strip()
            else:
                _log_tax_event("search_llm_started", **log_context, source=source)
                llm_started = time.perf_counter()
                subdivision = extract_subdivision_llm(property_record.get("legal1", ""))
                llm_ms = _elapsed_ms(llm_started)
                fuzzy = True
                _log_tax_event(
                    "search_llm_complete",
                    **log_context,
                    source=source,
                    llm_ms=llm_ms,
                    subdivision_found=bool(subdivision),
                    rss_mb=_current_rss_mb(),
                )
            if not subdivision:
                return jsonify(
                    {
                        "error": "Could not determine subdivision from property legal description",
                        "main_property": property_record,
                        "source": source,
                    }
                ), 422
        elif source == "liberty":
            subdivision = property_record.get("subdivision")
            if not subdivision or not subdivision_code:
                return jsonify(
                    {
                        "error": "Could not determine Liberty subdivision from tax data",
                        "main_property": property_record,
                        "source": source,
                    }
                ), 422
        elif source == "fort_bend":
            subdivision = property_record.get("subdivision")
            if not subdivision or not subdivision_code:
                return jsonify(
                    {
                        "error": "Could not determine Fort Bend neighborhood from tax data",
                        "main_property": property_record,
                        "source": source,
                    }
                ), 422
        else:
            legal_desc = property_record.get("legal1", "")
            if source == "chambers":
                _log_tax_event("search_llm_started", **log_context, source=source)
                llm_started = time.perf_counter()
                subdivision = extract_chambers_subdivision(legal_desc)
                llm_ms = _elapsed_ms(llm_started)
                subdivision_match_terms = build_chambers_subdivision_match_terms(
                    subdivision,
                    legal_desc,
                )
                _log_tax_event(
                    "search_llm_complete",
                    **log_context,
                    source=source,
                    llm_ms=llm_ms,
                    subdivision_found=bool(subdivision),
                    rss_mb=_current_rss_mb(),
                )
            else:
                _log_tax_event("search_llm_started", **log_context, source=source)
                llm_started = time.perf_counter()
                subdivision = extract_subdivision_llm(legal_desc)
                llm_ms = _elapsed_ms(llm_started)
                _log_tax_event(
                    "search_llm_complete",
                    **log_context,
                    source=source,
                    llm_ms=llm_ms,
                    subdivision_found=bool(subdivision),
                    rss_mb=_current_rss_mb(),
                )
            if not subdivision:
                return jsonify(
                    {
                        "error": "Could not extract subdivision from property legal description",
                        "main_property": property_record,
                        "source": source,
                    }
                ), 422

        _log_tax_event(
            "search_comparables_started",
            **log_context,
            source=source,
            limit=SEARCH_COMPARABLE_LIMIT,
            fuzzy=fuzzy,
            rss_mb=_current_rss_mb(),
        )
        comparables_started = time.perf_counter()
        comparables = find_comparables(
            subdivision,
            zip_code,
            market_value,
            source,
            main_sq_ft=main_sq_ft,
            fuzzy_subdivision=fuzzy,
            subdivision_code=subdivision_code,
            main_acreage=main_acreage,
            subdivision_match_terms=subdivision_match_terms,
            limit=SEARCH_COMPARABLE_LIMIT,
        )
        comparables_ms = _elapsed_ms(comparables_started)
        comparables_truncated = len(comparables) > SEARCH_COMPARABLE_LIMIT
        if comparables_truncated:
            comparables = comparables[:SEARCH_COMPARABLE_LIMIT]
        _log_tax_event(
            "search_comparables_complete",
            **log_context,
            source=source,
            comparables_ms=comparables_ms,
            comparables_returned=len(comparables),
            comparables_truncated=comparables_truncated,
            rss_mb=_current_rss_mb(),
        )

        cache_search_result(
            source=source,
            subdivision=subdivision,
            main_property_id=property_record["id"],
            contact_id=contact.id,
            zip_code=zip_code,
            neighborhood_code=neighborhood_code,
            subdivision_code=subdivision_code,
            main_sq_ft=main_sq_ft,
            main_acreage=main_acreage,
            fuzzy_subdivision=fuzzy,
            subdivision_match_terms=subdivision_match_terms,
        )

        _log_tax_event(
            "search_stats_started",
            **log_context,
            source=source,
            rss_mb=_current_rss_mb(),
        )
        stats_started = time.perf_counter()
        subdivision_stats = get_subdivision_stats(
            subdivision,
            zip_code,
            market_value,
            source,
            fuzzy_subdivision=fuzzy,
            subdivision_code=subdivision_code,
            subdivision_match_terms=subdivision_match_terms,
        )
        stats_ms = _elapsed_ms(stats_started)
        _log_tax_event(
            "search_stats_complete",
            **log_context,
            source=source,
            stats_ms=stats_ms,
            total_homes=(subdivision_stats or {}).get("total_homes"),
            rss_mb=_current_rss_mb(),
        )

        total_ms = _elapsed_ms(route_started)
        _log_tax_event(
            "search_complete",
            **log_context,
            source=source,
            total_ms=total_ms,
            lookup_ms=lookup_ms,
            llm_ms=llm_ms,
            comparables_ms=comparables_ms,
            stats_ms=stats_ms,
            comparables_returned=len(comparables),
            comparables_truncated=comparables_truncated,
            rss_mb=_current_rss_mb(),
        )

        return jsonify(
            {
                "source": source,
                "subdivision": subdivision,
                "main_property": property_record,
                "comparables": comparables,
                "total_comparables": len(comparables),
                "comparables_displayed": len(comparables),
                "comparables_truncated": comparables_truncated,
                "subdivision_stats": subdivision_stats,
            }
        )
    except HTTPException:
        raise
    except Exception:
        logger.exception(
            "tax_protest_search_failed contact_id=%s user_id=%s org_id=%s source=%s rss_mb=%s",
            contact.id,
            current_user.id,
            current_user.organization_id,
            source,
            _current_rss_mb(),
        )
        raise


@tax_protest_bp.route("/download-csv")
@login_required
@feature_required("TAX_PROTEST")
def download_csv():
    """Download CSV of comparables using cached search result."""
    started_at = time.perf_counter()
    export_data = _load_cached_export_data()
    output = StringIO()
    writer = csv.writer(output)

    headers = [
        "Type",
        "Address",
        "City",
        "Zip",
        "Market Value",
        "Sq Ft",
        "Acreage",
        "Subdivision",
        "Legal Description",
        "Account",
        "County",
    ]
    writer.writerow(headers)
    writer.writerows(export_data["rows"])

    output.seek(0)
    filename = _safe_filename(export_data["contact"].street_address, ".csv")
    _log_tax_event(
        "download_csv_complete",
        contact_id=export_data["contact"].id,
        user_id=current_user.id,
        org_id=current_user.organization_id,
        source=export_data["cached"].get("source"),
        rows=len(export_data["rows"]),
        total_ms=_elapsed_ms(started_at),
        rss_mb=_current_rss_mb(),
    )

    return Response(
        output.getvalue(),
        mimetype="text/csv",
        headers={
            "Content-Disposition": f"attachment; filename={filename}",
            "Content-Type": "text/csv",
        },
    )


@tax_protest_bp.route("/download-xlsx")
@login_required
@feature_required("TAX_PROTEST")
def download_xlsx():
    """Download an Excel report with a summary sheet and embedded chart."""
    started_at = time.perf_counter()
    export_data = _load_cached_export_data()
    _log_tax_event(
        "download_xlsx_started",
        contact_id=export_data["contact"].id,
        user_id=current_user.id,
        org_id=current_user.organization_id,
        source=export_data["cached"].get("source"),
        rows=len(export_data["rows"]),
        rss_mb=_current_rss_mb(),
    )
    workbook_stream = _build_xlsx_report(export_data)
    filename = _safe_filename(
        export_data["contact"].street_address, "_tax_protest_report.xlsx"
    )
    _log_tax_event(
        "download_xlsx_complete",
        contact_id=export_data["contact"].id,
        user_id=current_user.id,
        org_id=current_user.organization_id,
        source=export_data["cached"].get("source"),
        rows=len(export_data["rows"]),
        total_ms=_elapsed_ms(started_at),
        rss_mb=_current_rss_mb(),
    )

    return send_file(
        workbook_stream,
        mimetype=EXPORT_XLSX_MIMETYPE,
        as_attachment=True,
        download_name=filename,
    )

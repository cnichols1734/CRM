"""
Regression tests for the tax protest stats flow.
"""

from io import BytesIO
from types import SimpleNamespace
from zipfile import ZipFile

import feature_flags as feature_flags_module
import routes.tax_protest as tax_protest_route
import services.tax_protest_service as tax_protest_service
from openpyxl import load_workbook


class _QueryStub:
    def __init__(self, values):
        self._rows = [(value,) for value in values]

    def filter(self, *args, **kwargs):
        return self

    def with_entities(self, *args, **kwargs):
        return self

    def order_by(self, *args, **kwargs):
        return self

    def all(self):
        return list(self._rows)


class _RecordingQueryStub(_QueryStub):
    def __init__(self, values):
        super().__init__(values)
        self.filters = []

    def filter(self, *args, **kwargs):
        self.filters.extend(args)
        return self

    def join(self, *args, **kwargs):
        return self

    def group_by(self, *args, **kwargs):
        return self


def test_get_subdivision_stats_returns_list_distribution(monkeypatch):
    values = [390000, 420000, 450000, 610000, 615000, 700000, 820000]
    monkeypatch.setattr(
        tax_protest_service.LibertyProperty,
        "query",
        _QueryStub(values),
    )
    monkeypatch.setattr(
        tax_protest_service,
        "_find_liberty_sibling_codes",
        lambda subdivision_code: [],
    )

    stats = tax_protest_service.get_subdivision_stats(
        subdivision="LIBERTY OAKS",
        zip_code="77327",
        market_value=590000,
        source="liberty",
        subdivision_code="007206",
    )

    assert stats["total_homes"] == len(values)
    assert stats["min_value"] == min(values)
    assert stats["max_value"] == max(values)
    assert isinstance(stats["value_distribution"], list)
    assert sum(bucket["count"] for bucket in stats["value_distribution"]) == len(values)


def test_chambers_subdivision_stats_uses_home_filters(monkeypatch):
    query = _RecordingQueryStub([420000, 565000, 617950])
    monkeypatch.setattr(
        tax_protest_service.ChambersProperty,
        "query",
        query,
    )

    stats = tax_protest_service.get_subdivision_stats(
        subdivision="SELLERS STATION",
        zip_code="77523",
        market_value=565340,
        source="chambers",
    )

    filter_sql = " ".join(str(expr) for expr in query.filters)
    assert "improvement_hs_val" in filter_sql
    assert "improvement_nhs_val" in filter_sql
    assert "prop_street_number" in filter_sql
    assert "prop_street" in filter_sql
    assert stats["total_homes"] == 3


def test_build_chambers_subdivision_match_terms_broadens_truncated_name():
    terms = tax_protest_service.build_chambers_subdivision_match_terms(
        "PLANTATION ON CB",
        "LOT 31 SEC 5 PLANTATION ON CB",
    )

    assert "PLANTATION ON CB" in terms
    assert "PLANTATION ON" in terms


def test_extract_chambers_subdivision_prefers_llm_for_clean_name(monkeypatch):
    monkeypatch.setattr(
        tax_protest_service,
        "extract_subdivision_llm",
        lambda legal_description: "SELLERS STATION",
    )

    subdivision = tax_protest_service.extract_chambers_subdivision(
        "BK 1 LT 16 SELLERS STATION",
    )

    assert subdivision == "SELLERS STATION"


def test_extract_chambers_subdivision_fallback_strips_bk_prefix(monkeypatch):
    monkeypatch.setattr(
        tax_protest_service,
        "extract_subdivision_llm",
        lambda legal_description: None,
    )

    subdivision = tax_protest_service.extract_chambers_subdivision(
        "BK 1 LT 16 SELLERS STATION",
    )

    assert subdivision == "SELLERS STATION"


def test_extract_chambers_subdivision_fallback_strips_section_suffix(monkeypatch):
    monkeypatch.setattr(
        tax_protest_service,
        "extract_subdivision_llm",
        lambda legal_description: None,
    )

    subdivision = tax_protest_service.extract_chambers_subdivision(
        "BK 7 LT 8 LEGENDS BAY SEC 2",
    )

    assert subdivision == "LEGENDS BAY"


def test_search_property_returns_subdivision_stats(owner_a_client, monkeypatch):
    monkeypatch.setattr(feature_flags_module, "org_has_feature", lambda *args, **kwargs: True)

    fake_contact = SimpleNamespace(
        id=999,
        street_address="156 Maryville Lane",
        city="Cleveland",
        zip_code="77327",
    )
    fake_property = {
        "id": "liberty-1",
        "address": "156 Maryville Lane",
        "full_address": "156 Maryville Lane, Cleveland, TX 77327",
        "city": "Cleveland",
        "zip": "77327",
        "market_value": 591000,
        "sq_ft": 2184,
        "acreage": 0.23,
        "subdivision": "MARYVILLE",
        "subdivision_code": "007206",
        "neighborhood_code": None,
    }
    fake_stats = {
        "total_homes": 21,
        "lower_values": 2,
        "higher_values": 18,
        "percentile": 9.5,
        "value_distribution": [
            {"label": "$362k", "count": 7},
            {"label": "$462k", "count": 7},
            {"label": "$562k", "count": 2},
            {"label": "$662k", "count": 1},
            {"label": "$762k", "count": 0},
            {"label": "$862k", "count": 2},
            {"label": "$962k", "count": 0},
            {"label": "$1062k", "count": 2},
        ],
        "min_value": 362000,
        "max_value": 1162000,
        "median_value": 505000,
    }

    monkeypatch.setattr(
        tax_protest_route,
        "_authorized_contact",
        lambda contact_id: fake_contact,
    )
    monkeypatch.setattr(
        tax_protest_route,
        "find_property_in_tax_data",
        lambda street_address, city, zip_code: (fake_property, "liberty"),
    )
    monkeypatch.setattr(
        tax_protest_route,
        "find_comparables",
        lambda *args, **kwargs: [{"id": "comp-1", "market_value": 420000}],
    )
    monkeypatch.setattr(
        tax_protest_route,
        "cache_search_result",
        lambda **kwargs: None,
    )
    monkeypatch.setattr(
        tax_protest_route,
        "get_subdivision_stats",
        lambda *args, **kwargs: fake_stats,
    )

    response = owner_a_client.post(
        "/tax-protest/search",
        json={"contact_id": fake_contact.id},
    )

    assert response.status_code == 200
    payload = response.get_json()
    assert payload["subdivision"] == "MARYVILLE"
    assert payload["main_property"]["market_value"] == 591000
    assert payload["subdivision_stats"] == fake_stats


def test_search_property_expands_chambers_subdivision_match_terms(owner_a_client, monkeypatch):
    monkeypatch.setattr(feature_flags_module, "org_has_feature", lambda *args, **kwargs: True)

    fake_contact = SimpleNamespace(
        id=999,
        street_address="1602 Bayou Breeze",
        city="Cove",
        zip_code="77523",
    )
    fake_property = {
        "id": "chambers-1",
        "address": "1602 Bayou Breeze",
        "full_address": "1602 Bayou Breeze",
        "city": "Cove",
        "zip": "77523",
        "market_value": 530000,
        "sq_ft": 2526,
        "acreage": 1.53,
        "legal1": "LOT 31 SEC 5 PLANTATION ON CB",
        "legal2": None,
        "legal3": None,
        "legal4": None,
    }
    captured = {}

    monkeypatch.setattr(
        tax_protest_route,
        "_authorized_contact",
        lambda contact_id: fake_contact,
    )
    monkeypatch.setattr(
        tax_protest_route,
        "find_property_in_tax_data",
        lambda street_address, city, zip_code: (fake_property, "chambers"),
    )

    def fake_find_comparables(*args, **kwargs):
        captured["comparables_terms"] = kwargs.get("subdivision_match_terms")
        return []

    def fake_get_subdivision_stats(*args, **kwargs):
        captured["stats_terms"] = kwargs.get("subdivision_match_terms")
        return {
            "total_homes": 153,
            "lower_values": 0,
            "higher_values": 152,
            "percentile": 0.0,
            "value_distribution": [{"label": "$530k", "count": 153}],
            "min_value": 530000,
            "max_value": 900000,
            "median_value": 650000,
        }

    monkeypatch.setattr(tax_protest_route, "find_comparables", fake_find_comparables)
    monkeypatch.setattr(tax_protest_route, "get_subdivision_stats", fake_get_subdivision_stats)
    monkeypatch.setattr(tax_protest_route, "cache_search_result", lambda **kwargs: None)

    response = owner_a_client.post(
        "/tax-protest/search",
        json={"contact_id": fake_contact.id},
    )

    assert response.status_code == 200
    assert "PLANTATION ON" in captured["comparables_terms"]
    assert "PLANTATION ON" in captured["stats_terms"]


def test_tax_protest_search_contacts_marks_missing_address(owner_a_client, seed, monkeypatch):
    monkeypatch.setattr(feature_flags_module, "org_has_feature", lambda *args, **kwargs: True)

    # Own row: seed "Jane" is mutated by other tests (e.g. edit → "JaneEdited"), so do not rely on it.
    no_street = owner_a_client.post(
        "/contacts/create",
        data={
            "first_name": "TaxProtNoAddr",
            "last_name": "Search",
            "email": "taxprotnoaddr@test.com",
            "group_ids": str(seed["group_a1"]),
        },
        follow_redirects=True,
    )
    assert no_street.status_code == 200

    create_response = owner_a_client.post(
        "/contacts/create",
        data={
            "first_name": "Chris",
            "last_name": "Addressed",
            "street_address": "123 Main St",
            "city": "Houston",
            "state": "TX",
            "zip_code": "77001",
            "group_ids": str(seed["group_a1"]),
        },
        follow_redirects=True,
    )
    assert create_response.status_code == 200

    missing_response = owner_a_client.get("/tax-protest/search-contacts?q=TaxProtNoAddr")
    assert missing_response.status_code == 200
    missing_payload = missing_response.get_json()
    assert missing_payload, "no-address contact should appear in search (check create form validation)"
    assert missing_payload[0]["name"] == "TaxProtNoAddr Search"
    assert missing_payload[0]["has_address"] is False

    addressed_response = owner_a_client.get("/tax-protest/search-contacts?q=Addressed")
    assert addressed_response.status_code == 200
    addressed_payload = addressed_response.get_json()
    assert addressed_payload[0]["name"] == "Chris Addressed"
    assert addressed_payload[0]["has_address"] is True


def test_create_contact_ajax_returns_json_summary(owner_a_client, seed):
    response = owner_a_client.post(
        "/contacts/create",
        data={
            "first_name": "Tax",
            "last_name": "Modal",
            "street_address": "444 Cedar Lane",
            "city": "Houston",
            "state": "TX",
            "zip_code": "77002",
            "group_ids": str(seed["group_a1"]),
        },
        headers={"X-Requested-With": "XMLHttpRequest"},
    )

    assert response.status_code == 200
    payload = response.get_json()
    assert payload["status"] == "success"
    assert payload["contact"]["name"] == "Tax Modal"
    assert payload["contact"]["has_address"] is True
    assert "444 Cedar Lane" in payload["contact"]["address"]


def test_download_xlsx_returns_report_with_embedded_chart(owner_a_client, monkeypatch):
    monkeypatch.setattr(feature_flags_module, "org_has_feature", lambda *args, **kwargs: True)

    fake_contact = SimpleNamespace(
        id=999,
        street_address="156 Maryville Lane",
        city="Cleveland",
        zip_code="77327",
    )
    fake_cached = {
        "contact_id": fake_contact.id,
        "main_property_id": "liberty-1",
        "source": "liberty",
        "subdivision": "MARYVILLE",
        "zip_code": "77327",
        "main_sq_ft": 2184,
        "main_acreage": 0.23,
        "subdivision_code": "007206",
        "fuzzy_subdivision": False,
    }
    fake_main_property = {
        "id": "liberty-1",
        "address": "156 Maryville Lane",
        "full_address": "156 Maryville Lane, Cleveland, TX 77327",
        "city": "Cleveland",
        "zip": "77327",
        "market_value": 591000,
        "sq_ft": 2184,
        "acreage": 0.23,
        "legal1": "MARYVILLE SEC 1",
        "account": "007206000001",
    }
    fake_comparables = [
        {
            "id": "comp-1",
            "full_address": "144 Maryville Lane, Cleveland, TX 77327",
            "city": "Cleveland",
            "zip": "77327",
            "market_value": 420000,
            "sq_ft": 2030,
            "acreage": 0.22,
            "legal1": "MARYVILLE SEC 1",
            "account": "007206000002",
        },
        {
            "id": "comp-2",
            "full_address": "148 Maryville Lane, Cleveland, TX 77327",
            "city": "Cleveland",
            "zip": "77327",
            "market_value": 0,
            "sq_ft": 1900,
            "acreage": 0.20,
            "legal1": "MARYVILLE SEC 1",
            "account": "007206000003",
        }
    ]
    fake_stats = {
        "total_homes": 21,
        "lower_values": 2,
        "higher_values": 18,
        "percentile": 9.5,
        "value_distribution": [
            {"label": "$362k", "count": 7},
            {"label": "$462k", "count": 7},
            {"label": "$562k", "count": 2},
            {"label": "$662k", "count": 1},
            {"label": "$762k", "count": 0},
            {"label": "$862k", "count": 2},
            {"label": "$962k", "count": 0},
            {"label": "$1062k", "count": 2},
        ],
        "min_value": 362000,
        "max_value": 1162000,
        "median_value": 505000,
    }

    monkeypatch.setattr(
        tax_protest_route,
        "get_cached_search_result",
        lambda: fake_cached,
    )
    monkeypatch.setattr(
        tax_protest_route,
        "_authorized_contact",
        lambda contact_id: fake_contact,
    )
    monkeypatch.setattr(
        tax_protest_route,
        "get_main_property_by_id",
        lambda property_id, source: fake_main_property,
    )
    monkeypatch.setattr(
        tax_protest_route,
        "find_comparables",
        lambda *args, **kwargs: fake_comparables,
    )
    monkeypatch.setattr(
        tax_protest_route,
        "get_subdivision_stats",
        lambda *args, **kwargs: fake_stats,
    )

    response = owner_a_client.get("/tax-protest/download-xlsx")

    assert response.status_code == 200
    assert response.mimetype == (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    workbook = load_workbook(BytesIO(response.data))
    assert workbook.sheetnames == ["Summary", "Comparables", "Distribution"]
    assert workbook["Summary"]["A1"].value == "Tax Protest Neighborhood Report"
    assert workbook["Summary"]["F4"].value == 2
    assert workbook["Summary"]["B5"].value == 591000
    assert workbook["Summary"]["A6"].value == "# of homes valued for less"
    assert workbook["Summary"]["C6"].value == "# of homes valued for more"
    assert workbook["Comparables"]["A2"].value == "Subject Property"
    assert workbook["Comparables"]["A3"].value == "Comparable"
    assert workbook["Comparables"].max_row == 3

    with ZipFile(BytesIO(response.data)) as workbook_zip:
        media_files = [
            name for name in workbook_zip.namelist() if name.startswith("xl/media/")
        ]
    assert media_files
